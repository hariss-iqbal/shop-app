"""
Parse DigiKhata Bill Report PDF (layout-mode text) into structured Excel.
Uses pdftotext -layout output where columns are positionally aligned.

Column positions:
  ~col 3:  Row data lines (row#, date, bill#, inline detail, amount)
  ~col 47: Detail-only lines (customer name, products, notes)
"""

import csv
import re
import sys
import subprocess
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

# Product patterns (longest/most specific first)
PRODUCT_REGEXES = [
    (r'pixel\s*10\s*pro\s*xl', 'Google'),
    (r'pixel\s*10\s*pro', 'Google'),
    (r'pixel\s*10\b', 'Google'),
    (r'pixel\s*9\s*pro\s*xl', 'Google'),
    (r'pixel\s*9\s*pro', 'Google'),
    (r'pixel\s*9\b', 'Google'),
    (r'pixel\s*8a', 'Google'),
    (r'pixel\s*8\s*pro', 'Google'),
    (r'pixel\s*8\b', 'Google'),
    (r'pixel\s*7a', 'Google'),
    (r'pixel\s*7\s*pro', 'Google'),
    (r'pixel\s*7\b', 'Google'),
    (r'pixel\s*6a', 'Google'),
    (r'pixel\s*6\s*pro', 'Google'),
    (r'pixel\s*6\b', 'Google'),
    (r'pixel\s*4a\s*5g', 'Google'),
    (r'noorka\s*4a', 'Google'),
    (r's\s?26\s*ultra', 'Samsung'),
    (r's\s?25\s*ultra', 'Samsung'),
    (r's\s?25\s*plus', 'Samsung'),
    (r's\s?25\b', 'Samsung'),
    (r's\s?24\s*ultra', 'Samsung'),
    (r's\s?24\s*plus', 'Samsung'),
    (r's\s?24\b', 'Samsung'),
    (r's\s?23\s*ultra', 'Samsung'),
    (r's\s?23\s*plus', 'Samsung'),
    (r's\s?23\b', 'Samsung'),
    (r's\s?22\s*ultra', 'Samsung'),
    (r's\s?22\s*plus', 'Samsung'),
    (r's\s?22\b', 'Samsung'),
    (r's\s?21\s*ultra', 'Samsung'),
    (r's\s?21\s*plus', 'Samsung'),
    (r's\s?21\b', 'Samsung'),
    (r'fold\s*\d', 'Samsung'),
    (r'galaxy\s*\w', 'Samsung'),
    (r'a\d{2}s?\s*5g', 'Samsung'),
    (r'oneplus\s*13t', 'OnePlus'),
    (r'oneplus\s*13\b', 'OnePlus'),
    (r'oneplus\s*12', 'OnePlus'),
    (r'oneplus\s*\d', 'OnePlus'),
    (r'buds\s*pro\s*2', 'Google'),
    (r'buds\s*pro\s*1', 'Google'),
    (r'buds\s*pro\b', 'Google'),
    (r'buds\s*a\b', 'Google'),
]

BRAND_MAP = {
    'pixel': 'Google', 'buds pro': 'Google', 'buds a': 'Google',
    'noorka': 'Google', 'oneplus': 'OnePlus', 'fold': 'Samsung',
    'galaxy': 'Samsung',
}

NOISE_STRINGS = {
    'Start Using Digikhata Now', 'INSTALL', 'Help:',
    'Google STORE Lahore', '03214495590', 'Bill Report',
    'Total Bills', 'Total Amount', 'Report Generated',
    'Grand Total',
}

# Customer skip patterns
CUSTOMER_SKIP = [
    r'\d+\s*pcs', r'^\d+\s*days?\s*check', r'^brand new', r'^box pack',
    r'^official\b', r'^single sim', r'^MINT\b', r'^BRAND\b', r'^PSID',
    r'^CPID\b', r'^porcelain', r'^black\b', r'^white\b', r'^green\b',
    r'^blue\b', r'^voilet\b', r'^charger', r'^cable', r'^dual sim',
    r'^non pta', r'^\d+gb\b', r'^back cover', r'^check warrant',
    r'^battery\b', r'^software\b', r'^time over', r'^pcs$',
    r'^online\b', r'^physical\b', r'^pta\b',
]


def is_noise(line):
    s = line.strip()
    if not s:
        return True
    for ns in NOISE_STRINGS:
        if s.startswith(ns):
            return True
    if s.startswith('(As of Today'):
        return True
    if re.match(r'^\s*#\s+Date\s+Bill#', s):
        return True
    return False


def parse_amount(text):
    text = text.strip().replace(',', '')
    try:
        return int(text)
    except ValueError:
        return None


def extract_brand(text):
    lower = text.lower()
    for pattern, brand in BRAND_MAP.items():
        if pattern in lower:
            return brand
    if re.search(r'\bs\s?\d{2}', lower):
        return 'Samsung'
    return ''


def normalize_product(raw):
    return re.sub(r'\s*\d+\s*pcs\s*$', '', raw.strip(), flags=re.IGNORECASE).strip()


def extract_products_pcs(text):
    products = []
    for line in text.split('\n'):
        if re.search(r'\d+\s*pcs', line, re.IGNORECASE):
            qty_m = re.search(r'(\d+)\s*pcs', line, re.IGNORECASE)
            qty = int(qty_m.group(1)) if qty_m else 1
            name = normalize_product(line)
            if name and name.lower() not in ('pcs',):
                products.append({'name': name, 'qty': qty})
    return products


def extract_product_by_pattern(text):
    lower = text.lower()
    for pattern, brand in PRODUCT_REGEXES:
        m = re.search(pattern, lower)
        if m:
            return {'name': m.group(0).strip(), 'qty': 1, 'brand': brand}
    return None


def extract_imei(text):
    imeis = re.findall(r'\b(\d{15})\b', text)
    return ', '.join(imeis) if imeis else ''


def extract_storage(text):
    matches = re.findall(r'(\d+)\s*gb', text, re.IGNORECASE)
    vals = [int(m) for m in matches if int(m) in (32, 64, 128, 256, 512, 1024)]
    return f"{max(vals)}GB" if vals else ''


def extract_condition_notes(text):
    parts = []
    lower = text.lower()
    if 'brand new' in lower or 'box pack' in lower:
        parts.append('Brand new/Box pack')
    if 'official' in lower:
        parts.append('Official')
    if 'pta approved' in lower or 'online approved' in lower:
        parts.append('PTA Approved')
    if 'non pta' in lower:
        parts.append('Non-PTA')
    if 'cpid' in lower:
        parts.append('CPID')
    if 'refurbished' in lower:
        parts.append('Refurbished')
    wm = re.search(r'(\d+)\s*days?\s*check\s*warrant', lower)
    if wm:
        parts.append(f"{wm.group(1)} days check warranty")
    if 'exchange' in lower or 'adjusted' in lower:
        parts.append('Exchange/Adjustment')
    dm = re.search(r'(\d+)k?\s*discount', lower)
    if dm:
        parts.append(f"Discount: {dm.group(0)}")
    if 'remaining' in lower or 'balance' in lower:
        bm = re.search(r'(\d+)k?\s*remaining', lower)
        parts.append(f"Remaining balance: {bm.group(0)}" if bm else 'Has remaining balance')
    rm = re.search(r'(\d+)k?\s*(?:received|recieved)', lower)
    if rm:
        parts.append(f"Received: {rm.group(0)}")
    return '; '.join(parts)


def has_payment_info(text):
    lower = text.lower()
    return any(k in lower for k in ['remaining', 'balance', 'received', 'recieved', 'adjusted', 'exchange', 'paid'])


def extract_customer_name(detail_lines):
    for dl in detail_lines:
        dl = dl.strip()
        if not dl:
            continue
        if re.match(r'^\d{1,5}$', dl):
            continue
        skip = False
        for sp in CUSTOMER_SKIP:
            if re.search(sp, dl, re.IGNORECASE):
                skip = True
                break
        if skip:
            continue
        return dl
    return ''


def parse_layout_text(text_file):
    """
    Parse pdftotext -layout output.

    Entry anchor lines match:
      row#  date  bill#  [inline_detail]  [amount]
    Detail lines start at ~col 47 (high indentation).
    """
    with open(text_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Row data lines come in two formats:
    # With row#: "   7     17 Mar 26            208      detail text         47,000"
    # Without:   "         26 Jan 26   113               detail text         58,000"

    # Format 1: with row number
    row_with_num_re = re.compile(
        r'^\s{0,6}(\d{1,4})\s+'           # row number
        r'(\d{2}\s+\w{3}\s+\d{2})\s+'     # date
        r'(\d{1,4})'                       # bill number
        r'(.*?)$'                          # rest of line
    )

    # Format 2: without row number (entries 100+)
    row_no_num_re = re.compile(
        r'^\s{5,15}'                       # leading spaces (no row#)
        r'(\d{2}\s+\w{3}\s+\d{2})\s+'     # date
        r'(\d{1,4})'                       # bill number
        r'(.*?)$'                          # rest of line
    )

    # Detail-only lines (high indentation, ~col 30+)
    detail_re = re.compile(r'^\s{30,}(.+)$')

    # Standalone row number — may also have detail text after it at col ~35
    # e.g., "   10                              Adnan MZ Mobile"
    standalone_rownum_re = re.compile(r'^\s{0,6}(\d{1,4})\s*$')
    standalone_rownum_with_detail_re = re.compile(r'^\s{0,6}(\d{1,4})\s{10,}(.+)$')

    # ================================================================
    # TWO-PASS APPROACH
    # Pass 1: Classify every line as row-anchor, detail, standalone-rownum, or noise
    # Pass 2: Assign detail lines to entries based on position
    # ================================================================

    def _parse_rest(rest):
        """Extract amount and inline detail from the rest of a row line."""
        amount = None
        inline_detail = rest
        amt_match = re.search(r'\s{2,}([\d,]+)\s*$', rest)
        if amt_match:
            amt_val = parse_amount(amt_match.group(1))
            if amt_val and amt_val >= 100:
                amount = amt_val
                inline_detail = rest[:amt_match.start()].strip()
        elif re.match(r'^[\d,]+$', rest):
            amt_val = parse_amount(rest)
            if amt_val and amt_val >= 100:
                amount = amt_val
                inline_detail = ''
        return amount, inline_detail

    # Pass 1: classify lines — keep track of original line numbers and blank gaps
    classified = []  # list of (line_idx, type, data)
    # types: 'row', 'detail', 'gap'

    pending_row_parts = []  # accumulate split row number digits, e.g. ["37", "5"] → 375
    consecutive_blanks = 0

    for line_idx, raw_line in enumerate(lines):
        line = raw_line.rstrip()

        # Track blank lines for gap detection
        if not line.strip():
            consecutive_blanks += 1
            continue

        # Skip noise lines — but DON'T reset consecutive_blanks,
        # so page headers (noise) surrounded by blanks don't
        # create false gap markers at page boundaries
        if is_noise(line):
            continue

        # Emit a gap marker if we had 2+ blank lines (entry boundary).
        if consecutive_blanks >= 2 and classified and classified[-1][1] != 'gap':
            classified.append((line_idx, 'gap', None))
        consecutive_blanks = 0

        # Standalone row number (may include detail text)
        # e.g., "   10                              Adnan MZ Mobile"
        #
        # Row numbers for 100+ are split across multiple lines:
        #   "37"  + detail   (BEFORE row line)  → first part, goes to pending
        #   [ROW LINE]       → entry created with pending_row_parts
        #   "5"   + detail   (AFTER row line)   → trailing digit for LAST entry
        #
        # So: if we have a recent row entry and no pending parts yet,
        # this digit is a trailing part of the last entry's row number.
        # Otherwise, it's the start of the next entry's row number.
        sm_detail = standalone_rownum_with_detail_re.match(line)
        sm = standalone_rownum_re.match(line) if not sm_detail else None

        if (sm_detail or sm) and not detail_re.match(line):
            digit = sm_detail.group(1) if sm_detail else sm.group(1)

            # Is this a trailing digit for the last row entry?
            # Check: if classified has a recent 'row' and no pending parts yet
            if (not pending_row_parts and classified and
                    classified[-1][1] == 'row' and
                    len(digit) <= 1):
                # Append to last row entry's row_num
                last_row = classified[-1][2]
                if last_row['row_num'] is not None:
                    last_row['row_num'] = int(str(last_row['row_num']) + digit)
                else:
                    last_row['row_num'] = int(digit)
            else:
                # Start or continue building pending row number
                pending_row_parts.append(digit)

            # Capture detail text if present
            if sm_detail:
                detail_text = sm_detail.group(2).strip()
                if detail_text:
                    classified.append((line_idx, 'detail', detail_text))
            continue

        # Helper to build row number from accumulated parts + optional new part
        def _build_row_num(extra_part=None):
            parts = list(pending_row_parts)
            if extra_part is not None:
                parts.append(str(extra_part))
            if parts:
                return int(''.join(parts))
            return None

        # Format 1: with row number on the line
        m1 = row_with_num_re.match(line)
        if m1 and not detail_re.match(line):
            rest = m1.group(4).strip()
            amount, inline_detail = _parse_rest(rest)
            # The row number on this line is the first part; any pending parts
            # were from the previous entry's trailing digits — flush them
            classified.append((line_idx, 'row', {
                'row_num': int(m1.group(1)),
                'date': m1.group(2).strip(),
                'bill_num': int(m1.group(3)),
                'amount': amount,
                'inline_detail': inline_detail,
            }))
            pending_row_parts = []
            continue

        # Format 2: without row number (entries 100+)
        m2 = row_no_num_re.match(line)
        if m2 and not detail_re.match(line):
            rest = m2.group(3).strip()
            amount, inline_detail = _parse_rest(rest)
            # Use accumulated pending parts as the row number
            row_num = _build_row_num()
            classified.append((line_idx, 'row', {
                'row_num': row_num,
                'date': m2.group(1).strip(),
                'bill_num': int(m2.group(2)),
                'amount': amount,
                'inline_detail': inline_detail,
            }))
            pending_row_parts = []
            continue

        # Detail line
        dm = detail_re.match(line)
        if dm:
            detail_text = dm.group(1).strip()
            if detail_text:
                classified.append((line_idx, 'detail', detail_text))
            continue

        # Don't reset pending_row_parts on other content — the parts
        # accumulate across detail lines and row lines until consumed

    # Pass 2: Build entries by grouping detail lines around row anchors
    # Key insight from the PDF layout:
    #   [detail lines BEFORE row line] → belong to THIS entry (customer name, etc.)
    #   [row line]                     → entry anchor (date, bill#, amount, inline detail)
    #   [detail lines AFTER row line]  → belong to THIS entry (products, notes)
    #   [detail lines...]              → until next row line, then those before it = next entry
    #
    # So for each row anchor at position R:
    #   - Pre-details: detail lines between previous row's last post-detail and R
    #   - Post-details: detail lines between R and the next pre-detail boundary

    # ----------------------------------------------------------------
    # Pass 2: Build entries
    #
    # DigiKhata layout pattern per entry:
    #   [customer name]        ← detail BEFORE row (pre-row)
    #   [description lines]    ← detail BEFORE row (pre-row)
    #   [ROW LINE]             ← anchor with date, bill#, amount
    #   [more description]     ← detail AFTER row (post-row)
    #   [product N pcs]        ← detail AFTER row (LAST line of entry)
    #   --- gap / next entry starts ---
    #
    # Between two consecutive row anchors, ALL detail lines belong to
    # one of the two entries. The split point is the LAST "pcs" line —
    # everything up to and including it = prev entry's post-row details.
    # Everything after it = next entry's pre-row details.
    # ----------------------------------------------------------------

    row_indices = [i for i, (_, typ, _) in enumerate(classified) if typ == 'row']

    # For each pair of consecutive rows, collect details between them
    # and split at the LAST GAP marker (2+ blank lines = entry boundary).
    #
    # Between row[i] and row[i+1]:
    #   [post-row details for entry i] ... [GAP] ... [pre-row details for entry i+1]
    #
    # If multiple gaps exist, use the last one as the split.
    # If no gap, fall back to splitting at the last "pcs" line.

    # Collect items between each pair of rows
    def collect_between(start_ci, end_ci):
        """Collect detail lines and gap positions between two classified indices."""
        items = []  # list of ('detail', text) or ('gap', None)
        for k in range(start_ci, end_ci):
            _, typ, data = classified[k]
            if typ == 'detail':
                items.append(('detail', data))
            elif typ == 'gap':
                items.append(('gap', None))
        return items

    # Collect items before first row
    pre_first_items = collect_between(0, row_indices[0]) if row_indices else []
    pre_first = [text for typ, text in pre_first_items if typ == 'detail']

    # Build gaps between consecutive rows
    between = []
    for ri in range(len(row_indices)):
        start = row_indices[ri] + 1
        end = row_indices[ri + 1] if ri + 1 < len(row_indices) else len(classified)
        between.append(collect_between(start, end))

    def split_between_entries(items):
        """Split detail items between two row anchors into (prev_post, next_pre).

        The layout between two row lines:
          [post-row details: notes, product pcs lines]
          [GAP: 3 blank lines = entry boundary]
          [pre-row details: customer name, description, maybe more pcs lines]

        Strategy: find the FIRST gap that comes after ANY pcs line.
        That's the real entry boundary. Everything before = prev entry.
        Everything after = next entry's pre-details.
        """
        gap_positions = []
        seen_pcs = False
        first_gap_after_pcs = None

        for k, (typ, data) in enumerate(items):
            if typ == 'detail' and (
                re.search(r'\d+\s*pcs', data, re.IGNORECASE) or
                data.lower().strip() == 'pcs'
            ):
                seen_pcs = True
            elif typ == 'gap':
                gap_positions.append(k)
                if seen_pcs and first_gap_after_pcs is None:
                    first_gap_after_pcs = k

        # Use the first gap after any pcs line
        if first_gap_after_pcs is not None:
            before = [text for typ, text in items[:first_gap_after_pcs] if typ == 'detail']
            after = [text for typ, text in items[first_gap_after_pcs + 1:] if typ == 'detail']
            return before, after

        # No pcs seen — use last gap
        if gap_positions:
            gp = gap_positions[-1]
            before = [text for typ, text in items[:gp] if typ == 'detail']
            after = [text for typ, text in items[gp + 1:] if typ == 'detail']
            return before, after

        # Nothing to split on
        all_details = [text for typ, text in items if typ == 'detail']
        return all_details, []

    entries = []
    for ri, ci in enumerate(row_indices):
        row_data = classified[ci][2]

        # Pre-details for this entry
        if ri == 0:
            pre_details = pre_first
        else:
            _, pre_details = split_between_entries(between[ri - 1])

        # Post-details: split current gap to get just this entry's portion
        post_details, _ = split_between_entries(between[ri])

        # Build detail_lines: pre + inline + post
        inline = row_data.get('inline_detail', '')
        detail_lines = list(pre_details)
        if inline:
            detail_lines.append(inline)
        detail_lines.extend(post_details)

        # Check amounts in details
        amount = row_data['amount']
        if amount is None:
            for d in list(detail_lines):
                if re.match(r'^[\d,]+$', d):
                    amt = parse_amount(d)
                    if amt and amt >= 100:
                        amount = amt
                        detail_lines.remove(d)
                        break

        entries.append({
            'row_num': row_data['row_num'],
            'date': row_data['date'],
            'bill_num': row_data['bill_num'],
            'amount': amount,
            'detail_lines': detail_lines,
        })

    return entries


def load_contacts(contacts_csv):
    """Load contacts from CSV and build a name → phone lookup.

    Builds full name from: First Name + Middle Name + Last Name,
    and also uses "File As" field. Matches are case-insensitive and
    stripped of extra whitespace.
    """
    name_to_phone = {}

    with open(contacts_csv, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)

        for row in reader:
            if len(row) < 23:
                continue

            # Build full name from parts
            first = row[0].strip()
            middle = row[1].strip()
            last = row[2].strip()
            file_as = row[9].strip()
            phone = row[22].strip()

            if not phone:
                continue

            # Clean phone: take first number if multiple (split by :::)
            if ':::' in phone:
                phone = phone.split(':::')[0].strip()

            # Build name variants
            names = set()

            # Full name from parts
            full = ' '.join(p for p in [first, middle, last] if p)
            if full:
                names.add(full.lower())

            # File As field
            if file_as:
                names.add(file_as.lower())

            # First + Last (skip middle)
            if first and last:
                names.add(f"{first} {last}".lower())

            # Just first name if it's descriptive enough
            if first and len(first) > 3:
                names.add(first.lower())

            for name in names:
                if name and name not in name_to_phone:
                    name_to_phone[name] = phone

    return name_to_phone


def lookup_phone(customer_name, raw_details, contacts):
    """Look up phone number for a customer name.

    Tries multiple strategies:
    1. Exact match on customer name
    2. Exact match on full raw details joined as one line
    3. Progressively join raw detail lines (first 2, first 3, etc.)
    """
    if not customer_name:
        return ''

    # Strategy 1: exact match on customer name
    key = customer_name.strip().lower()
    if key in contacts:
        return contacts[key]

    # Strategy 2+3: try joining raw detail lines progressively
    # In DigiKhata contacts, the name is often the full detail text joined
    if raw_details:
        lines = [l.strip() for l in raw_details.split('\n') if l.strip()]
        # Try joining first N lines (2, 3, 4, ... up to all)
        for n in range(2, len(lines) + 1):
            joined = ' '.join(lines[:n]).lower()
            if joined in contacts:
                return contacts[joined]

    return ''


def process_entries(entries, contacts=None):
    """Convert raw entries to structured bill records."""
    bills = []
    for i, entry in enumerate(entries):
        details_text = '\n'.join(entry['detail_lines'])

        # Products
        products = extract_products_pcs(details_text)
        product_source = 'pcs_pattern' if products else None

        if not products:
            # Also check inline detail on the row line itself
            fallback = extract_product_by_pattern(details_text)
            if fallback:
                products = [{'name': fallback['name'], 'qty': fallback['qty']}]
                product_source = 'name_pattern'

        imei = extract_imei(details_text)
        storage = extract_storage(details_text)
        condition_notes = extract_condition_notes(details_text)
        customer_name = extract_customer_name(entry['detail_lines'])

        if products:
            product_summary = '; '.join(
                f"{p['name']} x{p['qty']}" if p['qty'] > 1 else p['name']
                for p in products
            )
            total_qty = sum(p['qty'] for p in products)
            primary_product = products[0]['name']
            brand = extract_brand(primary_product) or extract_brand(details_text)
        else:
            product_summary = ''
            total_qty = 0
            primary_product = ''
            brand = extract_brand(details_text)

        # Severity classification
        doubts = []
        severity = 'ok'

        if not entry['amount'] and not products:
            doubts.append('Empty entry — no amount or product')
            severity = 'error'
        elif not entry['amount']:
            doubts.append('No amount found')
            severity = 'error'
        elif not products:
            doubts.append('No product identified')
            severity = 'warning'

        if not customer_name:
            doubts.append('No customer name')
            if severity == 'ok':
                severity = 'warning'
        if product_source == 'name_pattern':
            doubts.append('Product from name pattern (no "N pcs")')
            if severity == 'ok':
                severity = 'info'
        if entry['amount'] and total_qty > 1 and len(products) == 1:
            doubts.append(f'Multi-qty ({total_qty} pcs) — per-unit price unknown')
            if severity in ('ok', 'info'):
                severity = 'warning'
        if has_payment_info(details_text):
            doubts.append('Has payment/balance info')
            if severity == 'ok':
                severity = 'info'
        if not brand and products:
            doubts.append('Brand unknown')
            if severity in ('ok', 'info'):
                severity = 'warning'

        # Phone lookup
        phone = lookup_phone(customer_name, details_text, contacts) if contacts else ''

        bills.append({
            'row_num': entry['row_num'],
            'date': entry['date'],
            'bill_num': entry['bill_num'],
            'customer_name': customer_name,
            'phone': phone,
            'brand': brand,
            'primary_product': primary_product,
            'product_summary': product_summary,
            'total_qty': total_qty,
            'amount': entry['amount'],
            'storage': storage,
            'imei': imei,
            'condition_notes': condition_notes,
            'raw_details': details_text,
            'doubt': severity != 'ok',
            'severity': severity,
            'doubt_reasons': '; '.join(doubts),
        })

    return bills


def write_excel(bills, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "DigiKhata Bills"

    headers = [
        'Row #', 'Date', 'Bill #', 'Customer Name', 'Phone', 'Phone Missing',
        'Brand', 'Primary Product', 'All Products', 'Qty', 'Amount (PKR)',
        'Storage', 'IMEI', 'Condition/Notes', 'Raw Details',
        'Severity', 'Doubt Reasons'
    ]

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    error_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    warning_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    info_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    severity_fonts = {
        'error': Font(bold=True, color='C62828'),
        'warning': Font(bold=True, color='E65100'),
        'info': Font(bold=True, color='1565C0'),
        'ok': Font(color='2E7D32'),
    }
    severity_fills = {
        'error': error_fill,
        'warning': warning_fill,
        'info': info_fill,
    }

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_idx, bill in enumerate(bills, 2):
        sev = bill['severity']
        phone_missing = 'YES' if bill['customer_name'] and not bill['phone'] else ''
        values = [
            bill['row_num'], bill['date'], bill['bill_num'],
            bill['customer_name'], bill['phone'], phone_missing,
            bill['brand'], bill['primary_product'], bill['product_summary'],
            bill['total_qty'] if bill['total_qty'] > 0 else '',
            bill['amount'], bill['storage'], bill['imei'],
            bill['condition_notes'], bill['raw_details'],
            sev.upper(), bill['doubt_reasons'],
        ]

        row_fill = severity_fills.get(sev)
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col in [8, 9, 14, 15, 17]))
            if row_fill:
                cell.fill = row_fill

        # Phone Missing highlight
        if phone_missing:
            pm_cell = ws.cell(row=row_idx, column=6)
            pm_cell.font = Font(bold=True, color='C62828')

        sev_cell = ws.cell(row=row_idx, column=16)
        sev_cell.font = severity_fonts.get(sev, Font())

    col_widths = {
        1: 8, 2: 13, 3: 8, 4: 30, 5: 18, 6: 14, 7: 12,
        8: 28, 9: 45, 10: 6, 11: 15, 12: 10,
        13: 20, 14: 40, 15: 50, 16: 10, 17: 45,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(bills) + 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    section_font = Font(bold=True, size=13, color='1F4E79')
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    ok_font = Font(size=11, color='2E7D32', bold=True)
    info_font = Font(size=11, color='1565C0', bold=True)
    warn_font = Font(size=11, color='E65100', bold=True)
    err_font = Font(size=11, color='C62828', bold=True)

    total_amount = sum(b['amount'] for b in bills if b['amount'])
    unique_customers = len(set(b['customer_name'] for b in bills if b['customer_name']))
    n_phone_matched = sum(1 for b in bills if b['phone'])
    n_phone_missing = sum(1 for b in bills if b['customer_name'] and not b['phone'])
    n_ok = sum(1 for b in bills if b['severity'] == 'ok')
    n_info = sum(1 for b in bills if b['severity'] == 'info')
    n_warning = sum(1 for b in bills if b['severity'] == 'warning')
    n_error = sum(1 for b in bills if b['severity'] == 'error')

    # Brand breakdown
    from collections import Counter
    brand_counts = Counter(b['brand'] for b in bills if b['brand'])
    # Top customers
    customer_counts = Counter(b['customer_name'] for b in bills if b['customer_name'])
    top_customers = customer_counts.most_common(15)
    # Monthly breakdown
    month_amounts = {}
    month_counts = {}
    for b in bills:
        if b['amount']:
            # Parse "09 Feb 26" → "Feb 26"
            parts = b['date'].split()
            if len(parts) == 3:
                month_key = f"{parts[1]} {parts[2]}"
                month_amounts[month_key] = month_amounts.get(month_key, 0) + b['amount']
                month_counts[month_key] = month_counts.get(month_key, 0) + 1

    # Doubt reasons
    doubt_counts = {}
    for b in bills:
        if b['doubt']:
            for r in b['doubt_reasons'].split('; '):
                r = r.strip()
                if r:
                    doubt_counts[r] = doubt_counts.get(r, 0) + 1

    ro = 1

    # Title
    ws2.cell(row=ro, column=1, value='DigiKhata Bill Report — Import Summary').font = Font(bold=True, size=15, color='1F4E79')
    ro += 2

    # Overview
    ws2.cell(row=ro, column=1, value='Overview').font = section_font
    ro += 1
    overview = [
        ('Total Bills', len(bills)),
        ('Total Amount (PKR)', f"{total_amount:,}"),
        ('Unique Customers', unique_customers),
        ('Phone Matched', n_phone_matched),
        ('Phone Missing', n_phone_missing),
        ('Date Range', f"{bills[-1]['date'] if bills else 'N/A'} to {bills[0]['date'] if bills else 'N/A'}"),
    ]
    for label, val in overview:
        ws2.cell(row=ro, column=1, value=label).font = label_font
        ws2.cell(row=ro, column=2, value=val).font = value_font
        ro += 1
    ro += 1

    # Severity
    ws2.cell(row=ro, column=1, value='Parsing Quality').font = section_font
    ro += 1
    sev_rows = [
        ('OK — ready to import', n_ok, f"{n_ok*100//len(bills)}%", ok_font),
        ('INFO — minor, review optional', n_info, f"{n_info*100//len(bills)}%", info_font),
        ('WARNING — review recommended', n_warning, f"{n_warning*100//len(bills)}%", warn_font),
        ('ERROR — needs manual fix', n_error, f"{n_error*100//len(bills)}%", err_font),
    ]
    ws2.cell(row=ro, column=1, value='Severity').font = label_font
    ws2.cell(row=ro, column=2, value='Count').font = label_font
    ws2.cell(row=ro, column=3, value='%').font = label_font
    ro += 1
    for label, count, pct, font in sev_rows:
        ws2.cell(row=ro, column=1, value=label).font = font
        ws2.cell(row=ro, column=2, value=count).font = font
        ws2.cell(row=ro, column=3, value=pct).font = font
        ro += 1
    ro += 1

    # Brand breakdown
    ws2.cell(row=ro, column=1, value='Brand Breakdown').font = section_font
    ro += 1
    ws2.cell(row=ro, column=1, value='Brand').font = label_font
    ws2.cell(row=ro, column=2, value='Bills').font = label_font
    ro += 1
    for brand, count in brand_counts.most_common():
        ws2.cell(row=ro, column=1, value=brand).font = value_font
        ws2.cell(row=ro, column=2, value=count).font = value_font
        ro += 1
    ro += 1

    # Monthly breakdown
    ws2.cell(row=ro, column=1, value='Monthly Breakdown').font = section_font
    ro += 1
    ws2.cell(row=ro, column=1, value='Month').font = label_font
    ws2.cell(row=ro, column=2, value='Bills').font = label_font
    ws2.cell(row=ro, column=3, value='Amount (PKR)').font = label_font
    ro += 1
    for month in sorted(month_amounts.keys(), key=lambda m: m, reverse=True):
        ws2.cell(row=ro, column=1, value=month).font = value_font
        ws2.cell(row=ro, column=2, value=month_counts[month]).font = value_font
        ws2.cell(row=ro, column=3, value=f"{month_amounts[month]:,}").font = value_font
        ro += 1
    ro += 1

    # Top customers
    ws2.cell(row=ro, column=1, value='Top 15 Customers').font = section_font
    ro += 1
    ws2.cell(row=ro, column=1, value='Customer').font = label_font
    ws2.cell(row=ro, column=2, value='Bills').font = label_font
    ro += 1
    for cust, count in top_customers:
        ws2.cell(row=ro, column=1, value=cust).font = value_font
        ws2.cell(row=ro, column=2, value=count).font = value_font
        ro += 1
    ro += 1

    # Doubt reasons
    ws2.cell(row=ro, column=1, value='Doubt Reasons Breakdown').font = section_font
    ro += 1
    ws2.cell(row=ro, column=1, value='Reason').font = label_font
    ws2.cell(row=ro, column=2, value='Count').font = label_font
    ro += 1
    for reason, count in sorted(doubt_counts.items(), key=lambda x: -x[1]):
        ws2.cell(row=ro, column=1, value=reason).font = value_font
        ws2.cell(row=ro, column=2, value=count).font = value_font
        ro += 1

    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 20

    wb.save(output_path)
    return len(bills), n_ok, n_info, n_warning, n_error


def main():
    pdf_file = '/Users/haris/Downloads/Bill_Report.pdf'
    contacts_file = '/Users/haris/Downloads/contacts.csv'
    layout_file = '/tmp/bill_report_layout.txt'
    output_file = '/Users/haris/Downloads/DigiKhata_Bills_Parsed.xlsx'

    # Extract with layout mode
    print("Extracting text from PDF (layout mode)...")
    subprocess.run(['pdftotext', '-layout', pdf_file, layout_file], check=True)

    print("Loading contacts...")
    contacts = load_contacts(contacts_file)
    print(f"  Loaded {len(contacts)} name→phone mappings")

    print(f"Parsing {layout_file}...")
    entries = parse_layout_text(layout_file)
    print(f"  Found {len(entries)} bill entries")

    print("Processing...")
    bills = process_entries(entries, contacts)

    # Phone match stats
    with_phone = sum(1 for b in bills if b['phone'])
    without_phone = sum(1 for b in bills if b['customer_name'] and not b['phone'])
    print(f"  Phone matched: {with_phone}")
    print(f"  Phone missing: {without_phone}")

    print(f"Writing Excel to {output_file}...")
    total, ok, info, warn, err = write_excel(bills, output_file)

    print(f"\nDone!")
    print(f"  Total:    {total}")
    print(f"  OK:       {ok}")
    print(f"  INFO:     {info}")
    print(f"  WARNING:  {warn}")
    print(f"  ERROR:    {err}")
    print(f"  Output:   {output_file}")


if __name__ == '__main__':
    main()
